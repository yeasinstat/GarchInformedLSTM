"""
GARCH-Informed LSTM (GINN) — Web App
======================================
Gradio app for Hugging Face Spaces.
Nav: Home | Model | Instructions | Developers
Model workflow: Data -> Summary Stats -> AR & GARCH -> LSTM Hyperparameters -> GINN
"""
import warnings, os, base64, time
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

import torch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import gradio as gr

import pipeline as P

np.random.seed(42)
torch.manual_seed(42)

# ═══════════════════════════════════════════════════════════
# Small shared helpers
# ═══════════════════════════════════════════════════════════

def df_to_html_table(df, max_height=None):
    """Render a DataFrame as plain HTML — avoids a Gradio Dataframe rendering
    quirk that can silently clip rows in wide/tall tables."""
    headers = "".join(f"<th style='padding:6px 10px;border:1px solid #ddd;background:#1F4E79;"
                       f"color:#fff;text-align:center;position:sticky;top:0'>{c}</th>" for c in df.columns)
    rows_html = ""
    for i, row in enumerate(df.itertuples(index=False)):
        bg = "#f5f7fa" if i % 2 == 0 else "#ffffff"
        cells = "".join(f"<td style='padding:6px 10px;border:1px solid #ddd;text-align:center'>{v}</td>" for v in row)
        rows_html += f"<tr style='background:{bg}'>{cells}</tr>"
    wrap_style = f"max-height:{max_height}px;overflow-y:auto;" if max_height else ""
    return (f"<div style='overflow-x:auto;{wrap_style}'><table style='border-collapse:collapse;width:100%;"
            f"font-family:monospace;font-size:0.85rem'><thead><tr>{headers}</tr></thead>"
            f"<tbody>{rows_html}</tbody></table></div>")


def load_file(file):
    if file is None:
        return None, gr.update(choices=[]), gr.update(choices=[]), "Upload a CSV or Excel file to begin."
    path = file.name if hasattr(file, "name") else file
    if str(path).lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path)
    cols = list(df.columns)
    msg = f"Loaded {len(df)} rows, {len(cols)} columns: {', '.join(cols)}"
    return df, gr.update(choices=cols, value=cols[0] if cols else None), \
           gr.update(choices=cols, value=cols[1] if len(cols) > 1 else None), msg


def summary_stats(df, time_col, value_col, stats_selected, show_line, show_box,
                   plot_width=8, plot_height=3.5, font_size=9):
    if df is None or value_col is None:
        return "Please upload data and select a study variable first.", None, None
    s = pd.to_numeric(df[value_col], errors="coerce").dropna()
    stat_map = {
        "Mean": s.mean(), "Median": s.median(), "Standard Deviation": s.std(),
        "Variance": s.var(), "Minimum": s.min(), "Maximum": s.max(),
        "Skewness": s.skew(), "Kurtosis": s.kurt(),
    }
    lines = [f"**{k}**: {stat_map[k]:.3f}" for k in stats_selected if k in stat_map]
    text = "\n\n".join(lines) if lines else "Select at least one statistic to display."

    line_fig = None
    if show_line:
        line_fig, ax = plt.subplots(figsize=(plot_width, plot_height))
        raw_x = df[time_col][:len(s)] if time_col in df.columns else pd.Series(range(len(s)))
        x_axis = None
        if not pd.api.types.is_numeric_dtype(raw_x):
            parsed = pd.to_datetime(raw_x, errors="coerce", dayfirst=True)
            # Guard against pandas mis-parsing plain numbers/short strings as
            # implausible dates (e.g. year 0000) — only accept a sane range.
            if parsed.notna().all() and parsed.dt.year.between(1678, 2262).all():
                x_axis = parsed
        if x_axis is not None:
            order = np.argsort(x_axis.values)
            ax.plot(x_axis.values[order], s.values[order], color="#1F4E79", linewidth=1.5)
            locator = mdates.AutoDateLocator(minticks=5, maxticks=12)
            ax.xaxis.set_major_locator(locator)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
            plt.setp(ax.get_xticklabels(), rotation=90, ha="center")
        else:
            x_pos = np.arange(len(s))
            ax.plot(x_pos, s.values, color="#1F4E79", linewidth=1.5)
            step = max(1, len(s) // 12)
            tick_idx = x_pos[::step]
            ax.set_xticks(tick_idx)
            ax.set_xticklabels([str(raw_x.iloc[i]) for i in tick_idx], rotation=90, ha="center")
        ax.tick_params(axis="both", labelsize=font_size)
        ax.set_title(f"{value_col} over time", fontsize=font_size + 2)
        ax.set_xlabel(time_col, fontsize=font_size); ax.set_ylabel(value_col, fontsize=font_size)
        line_fig.tight_layout()

    box_fig = None
    if show_box:
        box_fig, ax = plt.subplots(figsize=(plot_width, plot_height))
        ax.boxplot(s.values, patch_artist=True, boxprops=dict(facecolor="#D9E1F2"))
        ax.tick_params(axis="both", labelsize=font_size)
        ax.set_title(f"{value_col} — Boxplot", fontsize=font_size + 2)
        box_fig.tight_layout()

    return text, line_fig, box_fig


# ═══════════════════════════════════════════════════════════
# Step 3 — AR & GARCH
# ═══════════════════════════════════════════════════════════

def run_ar_garch(df, value_col, train_ratio, progress=gr.Progress()):
    if df is None or value_col is None:
        return "Upload data first.", None, None, None
    progress(0, desc="Computing returns...")
    data = pd.to_numeric(df[value_col], errors="coerce").dropna().values.astype(float)
    if len(data) < 20:
        return ("Need at least ~20 data points for meaningful AR/GARCH estimation "
                 "(returns + train/test split + sequence modeling).", None, None, None)

    returns = P.compute_returns(data)
    split = int(len(returns) * train_ratio)
    if len(returns) - split < 5:
        split = len(returns) - 5
    train_returns, test_returns = returns[:split], returns[split:]

    progress(0.1, desc="AR auto-tune...")
    best_lag, ar_model, ar_table = P.ar_autotune(train_returns)

    def garch_progress(frac, desc=""):
        progress(0.15 + 0.8 * frac, desc=desc)

    best_cfg, garch_fit, garch_table = P.garch_autotune(train_returns, best_lag, progress_fn=garch_progress)

    progress(0.97, desc="Computing ground-truth variance...")
    mu_train_fitted = ar_model.fittedvalues
    pad_len = len(train_returns) - len(mu_train_fitted)
    mu_train = np.concatenate([np.full(pad_len, mu_train_fitted[0]), mu_train_fitted])
    mu_test = np.asarray(ar_model.forecast(steps=len(test_returns)))

    sigma2_gt_train = (train_returns - mu_train) ** 2
    sigma2_gt_test = (test_returns - mu_test) ** 2

    sigma2_garch_train = P.fill_leading_nan((garch_fit.conditional_volatility / 100) ** 2)[:len(train_returns)]
    garch_fc = garch_fit.forecast(horizon=len(test_returns), reindex=False)
    sigma2_garch_test = P.fill_leading_nan(garch_fc.variance.values[-1] / 10000)[:len(test_returns)]

    garch_rmse = float(np.sqrt(np.mean((sigma2_gt_test - sigma2_garch_test) ** 2)))
    garch_mae = float(np.mean(np.abs(sigma2_gt_test - sigma2_garch_test)))

    state = dict(value_col=value_col, data=data, returns=returns, split=split,
                 train_returns=train_returns, test_returns=test_returns,
                 mu_train=mu_train, mu_test=mu_test,
                 sigma2_gt_train=sigma2_gt_train, sigma2_gt_test=sigma2_gt_test,
                 sigma2_garch_train=sigma2_garch_train, sigma2_garch_test=sigma2_garch_test,
                 best_lag=best_lag, best_cfg=best_cfg,
                 garch_params=dict(garch_fit.params), garch_aic=float(garch_fit.aic),
                 ar_aic=float(ar_model.aic), ar_table=ar_table, garch_table=garch_table,
                 garch_rmse=garch_rmse, garch_mae=garch_mae)

    p, q, mean, dist = best_cfg
    summary = (f"**Best AR order**: AR({best_lag})  (AIC={ar_model.aic:.3f})\n\n"
               f"**Best GARCH order**: GARCH({p},{q})  mean={mean}  dist={dist}  (AIC={garch_fit.aic:.3f})\n\n"
               f"**GARCH standalone variance RMSE**: {garch_rmse:.3f}\n\n"
               f"**GARCH standalone variance MAE**: {garch_mae:.3f}")

    fig, ax1 = plt.subplots(figsize=(9, 3.5))
    ax1.plot(returns, color="#3498DB", lw=1.3)
    ax1.axvline(x=split, color="red", ls="--", lw=1.2, label="Train / Test split")
    ax1.axhline(0, color="black", lw=0.7, ls=":")
    ax1.set_xlabel("Time Step"); ax1.set_ylabel("Return")
    ax1.legend(fontsize=8)
    fig.tight_layout()

    # GARCH parameter table — Estimate, Standard Error, t value, p-value
    param_rows = []
    for k in garch_fit.params.index:
        est = float(garch_fit.params[k])
        se = float(garch_fit.std_err[k])
        tval = float(garch_fit.tvalues[k])
        pval = float(garch_fit.pvalues[k])
        p_display = "< 0.0001*" if pval < 0.0001 else f"{pval:.3f}"
        param_rows.append([k, f"{est:.3f}", f"{se:.3f}", f"{tval:.3f}", p_display])
    param_df = pd.DataFrame(param_rows, columns=["Parameter", "Estimate", "Standard Error", "t value", "p-value"])
    param_html = df_to_html_table(param_df)

    # ── Excel export: order/fit summary + full GARCH parameter table ──
    summary_rows = [
        ["Best AR order", f"AR({best_lag})"],
        ["AR AIC", round(float(ar_model.aic), 4)],
        ["Best GARCH order", f"GARCH({p},{q}) mean={mean} dist={dist}"],
        ["GARCH AIC", round(float(garch_fit.aic), 4)],
        ["GARCH standalone variance RMSE", round(garch_rmse, 3)],
        ["GARCH standalone variance MAE", round(garch_mae, 3)],
        ["Train / Test split ratio", round(float(train_ratio), 2)],
    ]
    summary_out_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
    os.makedirs(out_dir, exist_ok=True)
    ar_garch_out_path = os.path.join(out_dir, "AR_GARCH_Results.xlsx")
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")

    def _write_sheet(ws, df_):
        for ci, h in enumerate(df_.columns, 1):
            c = ws.cell(1, ci, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
            c.alignment = Alignment(horizontal="center"); c.border = border
        for ri, row in enumerate(df_.itertuples(index=False), 2):
            for ci, v in enumerate(row, 1):
                c = ws.cell(ri, ci, v); c.border = border; c.alignment = Alignment(horizontal="center")
        for ci in range(1, len(df_.columns) + 1):
            ws.column_dimensions[chr(64 + ci) if ci <= 26 else "A"].width = 22

    _write_sheet(wb.create_sheet("Summary"), summary_out_df)
    _write_sheet(wb.create_sheet("GARCH_Parameters"), param_df)
    wb.save(ar_garch_out_path)

    progress(1.0, desc="Done")
    return summary, fig, state, param_html, ar_garch_out_path


# ═══════════════════════════════════════════════════════════
# Step 4 — LSTM Hyperparameters
# ═══════════════════════════════════════════════════════════

def run_lstm_tuning(ar_garch_state, seq_len, tune_mode,
                     hidden, layers, lr, dropout,
                     g_hidden_min, g_hidden_max, g_hidden_step,
                     g_layers_min, g_layers_max, g_layers_step,
                     g_lr_min, g_lr_max, g_lr_step,
                     g_dropout_min, g_dropout_max, g_dropout_step,
                     b_hidden_min, b_hidden_max,
                     b_layers_min, b_layers_max,
                     b_lr_min, b_lr_max,
                     b_dropout_min, b_dropout_max,
                     n_calls,
                     progress=gr.Progress()):
    if ar_garch_state is None:
        return "Complete the AR & GARCH step first.", None
    np.random.seed(42); torch.manual_seed(42)
    progress(0, desc="Preparing sequences...")
    try:
        seqs = P.prepare_sequences(ar_garch_state, int(seq_len))
    except ValueError as e:
        return f"⚠️ {e}", None

    train_loader, Xv, yv, scaler = seqs["train_loader"], seqs["X_val_t"], seqs["y_val_t"], seqs["scaler"]

    if tune_mode == "Manual":
        progress(0.5, desc="Applying manual hyperparameters...")
        best_hp = dict(hidden=int(hidden), layers=int(layers), lr=float(lr), dropout=float(dropout))
        progress(1.0, desc="Done")
        msg = (f"**Manual hyperparameters set** — hidden={best_hp['hidden']}, "
               f"layers={best_hp['layers']}, lr={best_hp['lr']}, dropout={best_hp['dropout']}")

    elif tune_mode == "Grid Search":
        param_grid = {
            "hidden_size": P.build_grid_values(g_hidden_min, g_hidden_max, g_hidden_step, True),
            "num_layers": P.build_grid_values(g_layers_min, g_layers_max, g_layers_step, True),
            "lr": P.build_grid_values(g_lr_min, g_lr_max, g_lr_step, False),
            "dropout": P.build_grid_values(g_dropout_min, g_dropout_max, g_dropout_step, False),
        }
        n_combos = (len(param_grid["hidden_size"]) * len(param_grid["num_layers"]) *
                    len(param_grid["lr"]) * len(param_grid["dropout"]))
        best_hp, _ = P.grid_search_lstm(param_grid, train_loader, Xv, yv, scaler, progress_fn=progress)
        msg = (f"**Grid search complete** ({n_combos} combinations tried) — "
               f"best: hidden={best_hp['hidden']}, layers={best_hp['layers']}, "
               f"lr={best_hp['lr']:.4f}, dropout={best_hp['dropout']:.2f} (val RMSE={best_hp['rmse']:.6f})")

    else:  # Bayesian Search
        best_hp, _ = P.bayes_search_lstm(
            train_loader, Xv, yv, scaler,
            hidden_range=(int(b_hidden_min), int(b_hidden_max)),
            layers_range=(int(b_layers_min), int(b_layers_max)),
            lr_range=(float(b_lr_min), float(b_lr_max)),
            dropout_range=(float(b_dropout_min), float(b_dropout_max)),
            n_calls=int(n_calls), progress_fn=progress)
        msg = (f"**Bayesian optimization complete** (Gaussian Process, {int(n_calls)} evaluations) — "
               f"best: hidden={best_hp['hidden']}, layers={best_hp['layers']}, "
               f"lr={best_hp['lr']:.4f}, dropout={best_hp['dropout']:.2f} (val RMSE={best_hp['rmse']:.6f})")

    return msg, best_hp


# ═══════════════════════════════════════════════════════════
# Step 5 — GINN run
# ═══════════════════════════════════════════════════════════

def run_ginn_pipeline(ar_garch_state, seq_len, best_hp, lambdas_str, epochs, patience,
                       progress=gr.Progress()):
    if ar_garch_state is None or best_hp is None:
        return ("Complete the AR & GARCH and LSTM Hyperparameters steps first.",
                 None, None, None, None, None)

    progress(0, desc="Preparing sequences...")
    try:
        seqs = P.prepare_sequences(ar_garch_state, int(seq_len))
    except ValueError as e:
        return f"⚠️ {e}", None, None, None, None, None

    # lambda=0.0 (Standard LSTM baseline) is always included
    user_lams = sorted(set(float(x.strip()) for x in lambdas_str.split(",") if x.strip() != ""))
    lambdas = sorted(set([0.0] + user_lams))

    train_loader = seqs["train_loader"]
    Xv, yv = seqs["X_val_t"], seqs["y_val_t"]
    Xte, yte = seqs["X_test_t"], seqs["y_test_t"]
    Xtr_full, ytr_full = seqs["X_train_full_t"], seqs["y_train_full_t"]
    garch_tr_t = seqs["garch_tr_t"]
    scaler = seqs["scaler"]

    rows = []
    all_preds = {}
    rmse_std = None
    for i, lam in enumerate(lambdas):
        progress(i / len(lambdas), desc=f"Training λ={lam} ({i + 1}/{len(lambdas)})...")
        np.random.seed(42); torch.manual_seed(42)
        t0 = time.time()
        model, pr, ac, rmse, mae, smape, hist, npar = P.train_ginn(
            best_hp, train_loader, Xv, yv, Xte, yte, scaler, garch_tr_t, lam,
            epochs=int(epochs), patience=int(patience),
            X_tr=seqs["X_tr_t"], y_tr=seqs["y_tr_t"])
        elapsed = time.time() - t0
        tr_pr, tr_ac, tr_rmse, tr_mae, tr_smape = P.evaluate_variance(model, Xtr_full, ytr_full, scaler)
        label = "Standard (λ=0.0)" if lam == 0.0 else f"λ={lam}"
        if lam == 0.0:
            rmse_std = rmse
        rows.append([label, round(tr_rmse, 3), round(tr_mae, 3), round(tr_smape, 3),
                     round(rmse, 3), round(mae, 3), round(smape, 3), npar, round(elapsed, 2)])
        all_preds[label] = dict(train=tr_pr, test=pr)

    # Improvement vs Standard, now that rmse_std is known
    for row in rows:
        rmse_val = row[4]
        row.append(round((rmse_std - rmse_val) / rmse_std * 100, 2) if rmse_std else None)

    results_df = pd.DataFrame(rows, columns=["Model", "Train RMSE", "Train MAE", "Train SMAPE (%)",
                                              "Test RMSE", "Test MAE", "Test SMAPE (%)", "Params",
                                              "Time (s)", "Improvement vs Standard (%)"])

    train_actual_inv = scaler.inverse_transform(seqs["y_tr_t"].numpy().reshape(-1, 1)).flatten()
    # (kept for reference; full train actual below uses the full train tensor)
    train_actual_full_inv = scaler.inverse_transform(ytr_full.numpy().reshape(-1, 1)).flatten()
    test_actual_inv = scaler.inverse_transform(yte.numpy().reshape(-1, 1)).flatten()

    train_table = {"Time Step": np.arange(1, len(train_actual_full_inv) + 1),
                    "Actual": np.round(train_actual_full_inv, 8)}
    for label, preds in all_preds.items():
        train_table[label] = np.round(preds["train"], 8)
    train_pred_df = pd.DataFrame(train_table)

    test_table = {"Time Step": np.arange(1, len(test_actual_inv) + 1),
                   "Actual": np.round(test_actual_inv, 8)}
    for label, preds in all_preds.items():
        test_table[label] = np.round(preds["test"], 8)
    test_pred_df = pd.DataFrame(test_table)

    p, q, mean, dist = ar_garch_state["best_cfg"]
    config_rows = [
        ["Sequence Length (lag window)", int(seq_len)],
        ["AR Order", f"AR({ar_garch_state['best_lag']})"],
        ["GARCH Order", f"GARCH({p},{q}) mean={mean} dist={dist}"],
        ["LSTM Hidden Size", int(best_hp["hidden"])],
        ["LSTM Num Layers", int(best_hp["layers"])],
        ["LSTM Learning Rate", float(best_hp["lr"])],
        ["LSTM Dropout", float(best_hp["dropout"])],
        ["λ Values Tested", ", ".join(f"{l:g}" for l in lambdas)],
        ["Max Epochs", int(epochs)],
        ["Early-Stopping Patience", int(patience)],
    ]
    config_df = pd.DataFrame(config_rows, columns=["Parameter", "Value"])

    # ── Excel export ──
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "GINN_Results.xlsx")
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")

    def _write_sheet(ws, df_):
        for ci, h in enumerate(df_.columns, 1):
            c = ws.cell(1, ci, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
            c.alignment = Alignment(horizontal="center"); c.border = border
        for ri, row in enumerate(df_.itertuples(index=False), 2):
            for ci, v in enumerate(row, 1):
                c = ws.cell(ri, ci, v); c.border = border; c.alignment = Alignment(horizontal="center")
        for ci in range(1, len(df_.columns) + 1):
            ws.column_dimensions[chr(64 + ci) if ci <= 26 else "A"].width = 15

    _write_sheet(wb.create_sheet("Summary"), results_df)
    _write_sheet(wb.create_sheet("Config"), config_df)

    ar_df = pd.DataFrame(ar_garch_state["ar_table"])
    garch_df = pd.DataFrame(ar_garch_state["garch_table"])
    garch_params_df = pd.DataFrame([[k, round(float(v), 6)] for k, v in ar_garch_state["garch_params"].items()],
                                    columns=["Parameter", "Estimate"])
    _write_sheet(wb.create_sheet("AR_Tuning"), ar_df)
    _write_sheet(wb.create_sheet("GARCH_Tuning"), garch_df)
    _write_sheet(wb.create_sheet("GARCH_Params"), garch_params_df)
    _write_sheet(wb.create_sheet("Train_Predictions"), train_pred_df)
    _write_sheet(wb.create_sheet("Test_Predictions"), test_pred_df)
    wb.save(out_path)

    progress(1.0, desc="Done")
    return ("Training complete for all λ values.", df_to_html_table(results_df),
             out_path, df_to_html_table(train_pred_df, max_height=400),
             df_to_html_table(test_pred_df, max_height=400), config_df)


# ═══════════════════════════════════════════════════════════
# Developers page (shared design)
# ═══════════════════════════════════════════════════════════

ASSETS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")

DEVELOPERS = [
    dict(name="Ranjit Kumar Paul", role="National Fellow",
         affiliation=["ICAR-Indian Agricultural Statistics Research Institute",
                      "Library Avenue, Pusa, New Delhi, Delhi – 110012, India"],
         email="ranjitstat@gmail.com",
         scholar="https://scholar.google.com/citations?hl=en&user=wBWuZJgAAAAJ&view_op=list_works&sortby=pubdate",
         photo="ranjit_kumar_paul.png"),
    dict(name="Md Yeasin", role="Scientist",
         affiliation=["Division of Statistical Ecology and Environmental Statistics",
                      "ICAR-Indian Agricultural Statistics Research Institute",
                      "Library Avenue, Pusa, New Delhi, Delhi – 110012, India"],
         email="yeasin.iasri@gmail.com",
         scholar="https://scholar.google.com/citations?user=xejMKD0AAAAJ",
         photo="md_yeasin.png"),
    dict(name="Pushkar Bora", role="Researcher",
         affiliation=["Discipline of Agricultural Statistics",
                      "ICAR-Indian Agricultural Statistics Research Institute",
                      "Library Avenue, Pusa, New Delhi, Delhi – 110012, India"],
         email="borapushkar1999@gmail.com",
         scholar="https://scholar.google.com/citations?user=xVIXGlwAAAAJ&hl=en",
         photo="pushkar_bora.jpg"),
    dict(name="Manojit Mandal", role="MSc Student",
         affiliation=["Discipline of Agricultural Statistics",
                      "ICAR-Indian Agricultural Statistics Research Institute",
                      "Library Avenue, Pusa, New Delhi, Delhi – 110012, India"],
         email=None,
         scholar=None,
         photo="manojit.jpeg"),
]

COUNTER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "visitor_count.txt")

def get_and_increment_visitor_count():
    try:
        with open(COUNTER_FILE, "r") as f:
            count = int(f.read().strip() or "0")
    except (FileNotFoundError, ValueError):
        count = 0
    count += 1
    try:
        with open(COUNTER_FILE, "w") as f:
            f.write(str(count))
    except OSError:
        pass
    return count
def _img_to_data_uri(filename):
    path = os.path.join(ASSETS_DIR, filename)
    ext = os.path.splitext(filename)[1].lower()
    mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
    try:
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")
        return f"data:{mime};base64,{b64}"
    except FileNotFoundError:
        return ""


def build_developers_html():
    cards = []
    for dev in DEVELOPERS:
        img_src = _img_to_data_uri(dev["photo"])
        affil_html = "".join(f"<div class='dev-affil'>{line}</div>" for line in dev["affiliation"])
        links = ""
        if dev.get("email"):
            links += f'<a class="dev-btn dev-btn-email" href="mailto:{dev["email"]}">✉&nbsp;Email</a>'
        if dev.get("scholar"):
            links += f'<a class="dev-btn dev-btn-scholar" href="{dev["scholar"]}" target="_blank" rel="noopener">🎓&nbsp;Google Scholar</a>'
        links_html = f'<div class="dev-links">{links}</div>' if links else ""
        cards.append(f"""
        <div class="dev-card">
            <div class="dev-photo-wrap"><img class="dev-photo" src="{img_src}" alt="{dev['name']}"/></div>
            <div class="dev-info">
                <div class="dev-name">{dev['name']}</div>
                <div class="dev-role">{dev['role']}</div>
                {affil_html}
                {links_html}
            </div>
        </div>""")
    style = """
    <style>
    .dev-wrap { max-width: 640px; margin: 4px auto 20px auto; font-family: inherit; }
    .dev-wrap .dev-intro { color: #374151 !important; font-size: 0.97rem; line-height: 1.5; margin-bottom: 18px; }
    .dev-wrap .dev-grid { display: flex; flex-direction: column; gap: 16px; }
    .dev-wrap .dev-card {
        display: flex; align-items: center; gap: 18px;
        background: linear-gradient(145deg, #0b1120, #131c33) !important;
        border: 1px solid #263352 !important; border-radius: 16px !important;
        padding: 18px 22px !important; box-shadow: 0 4px 18px rgba(0,0,0,0.28) !important;
        transition: transform 0.15s ease, box-shadow 0.15s ease;
    }
    .dev-wrap .dev-card:hover { transform: translateY(-2px); box-shadow: 0 8px 26px rgba(0,0,0,0.38) !important; }
    .dev-wrap .dev-photo-wrap { flex-shrink: 0; padding: 2px; border-radius: 14px !important;
        background: linear-gradient(135deg, #22d3ee, #6d28d9) !important; }
    .dev-wrap .dev-photo { display: block; width: 78px !important; height: 78px !important;
        object-fit: cover !important; border-radius: 12px !important; border: 2px solid #0b1120 !important; }
    .dev-wrap .dev-info { display: flex; flex-direction: column; gap: 2px; min-width: 0; }
    .dev-wrap .dev-name { color: #f8fafc !important; font-size: 1.12rem !important; font-weight: 700 !important; line-height: 1.25; }
    .dev-wrap .dev-role { color: #22d3ee !important; font-weight: 600 !important; font-size: 0.88rem !important; margin-bottom: 3px; }
    .dev-wrap .dev-affil { color: #94a3b8 !important; font-size: 0.8rem !important; line-height: 1.3; }
    .dev-wrap .dev-links { display: flex; gap: 8px; margin-top: 10px; flex-wrap: wrap; }
    .dev-wrap .dev-btn { display: inline-flex !important; align-items: center; gap: 5px; padding: 5px 13px !important;
        border-radius: 999px !important; font-size: 0.78rem !important; font-weight: 600 !important;
        text-decoration: none !important; color: #ffffff !important; border: none !important; }
    .dev-wrap .dev-btn-email { background: #0e7490 !important; }
    .dev-wrap .dev-btn-email:hover { background: #06b6d4 !important; }
    .dev-wrap .dev-btn-scholar { background: #6d28d9 !important; }
    .dev-wrap .dev-btn-scholar:hover { background: #8b5cf6 !important; }
    </style>
    """
    return style + f'<div class="dev-wrap"><div class="dev-grid">' + "".join(cards) + "</div></div>"


# ═══════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════

STAT_CHOICES = ["Mean", "Median", "Standard Deviation", "Variance",
                 "Minimum", "Maximum", "Skewness", "Kurtosis"]

FONT_HEAD = """
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
"""

CUSTOM_THEME = gr.themes.Base(
    primary_hue=gr.themes.colors.orange,
    secondary_hue=gr.themes.colors.violet,
    neutral_hue=gr.themes.colors.slate,
    font=[gr.themes.GoogleFont("Inter"), "sans-serif"],
    font_mono=[gr.themes.GoogleFont("JetBrains Mono"), "monospace"],
).set(
    body_background_fill="#faf6ee", body_background_fill_dark="#faf6ee",
    background_fill_primary="#ffffff", background_fill_primary_dark="#ffffff",
    background_fill_secondary="#fdf7ea", background_fill_secondary_dark="#fdf7ea",
    border_color_primary="#ecdfc4", border_color_primary_dark="#ecdfc4",
    block_background_fill="#ffffff", block_background_fill_dark="#ffffff",
    block_border_color="#ecdfc4", block_border_color_dark="#ecdfc4",
    block_label_background_fill="#fdf1de", block_label_background_fill_dark="#fdf1de",
    block_title_text_color="#b5591a", block_title_text_color_dark="#b5591a",
    body_text_color="#2a2418", body_text_color_dark="#2a2418",
    body_text_color_subdued="#8a7f68", body_text_color_subdued_dark="#8a7f68",
    input_background_fill="#fffdf8", input_background_fill_dark="#fffdf8",
    input_border_color="#ecdfc4", input_border_color_dark="#ecdfc4",
    button_primary_background_fill="linear-gradient(135deg, #ff9d4d, #e2725b)",
    button_primary_background_fill_dark="linear-gradient(135deg, #ff9d4d, #e2725b)",
    button_primary_text_color="#ffffff", button_primary_text_color_dark="#ffffff",
    button_secondary_background_fill="#ffffff", button_secondary_background_fill_dark="#ffffff",
    button_secondary_text_color="#2a2418", button_secondary_text_color_dark="#2a2418",
    button_secondary_border_color="#ecdfc4", button_secondary_border_color_dark="#ecdfc4",
    shadow_drop="0 10px 30px rgba(180,120,40,.14)",
)

CUSTOM_CSS = """
:root {
  --coral: #e2725b; --amber: #f5a742; --teal: #2bb8a8; --violet: #8b5cf6; --ink: #2a2418;
}
@keyframes drift {
  0%   { transform: translate(0,0) scale(1); }
  50%  { transform: translate(3%,-4%) scale(1.06); }
  100% { transform: translate(0,0) scale(1); }
}
.gradio-container {
  font-family: 'Inter', sans-serif !important;
  position: relative; background: #faf6ee !important; overflow-x: hidden;
}
.gradio-container::before {
  content: ''; position: fixed; inset: -10%; z-index: 0; pointer-events: none;
  background:
    radial-gradient(38% 30% at 12% 8%,  rgba(245,167,66,.35), transparent 70%),
    radial-gradient(35% 28% at 88% 6%,  rgba(139,92,246,.28), transparent 70%),
    radial-gradient(40% 32% at 90% 85%, rgba(43,184,168,.30), transparent 70%),
    radial-gradient(35% 30% at 8% 90%,  rgba(226,114,91,.28), transparent 70%);
  filter: blur(40px); animation: drift 22s ease-in-out infinite;
}
.gradio-container > * { position: relative; z-index: 1; }
h1, h2, h3, .prose h1, .prose h2, .prose h3 {
  font-family: 'Fraunces', serif !important; font-weight: 600 !important; letter-spacing: -.01em;
  color: var(--ink) !important;
}
#nav-title {
  margin: -16px -16px 0 -16px !important;
  padding: 0 !important;
  line-height: 0;
}
#nav-title img {
  display: block !important;
  width: 100% !important;
  height: auto !important;
  max-height: none !important;
}

/* top navigation bar */
.tabs > .tab-nav {
  display: flex !important; align-items: center; justify-content: flex-start;
  width: 100% !important; max-width: 100% !important;
  background: linear-gradient(90deg, #4a2545, #6b3a63) !important;
  border: none !important; border-radius: 0 !important;
  padding: 14px 28px !important; margin: -16px -16px 0 -16px !important;
  box-shadow: 0 4px 20px rgba(74,37,69,.25);
}
.tabs > .tab-nav::before {
  content: '📉  GARCH-GUIDED LSTM'; font-family: 'Fraunces', serif; font-weight: 700;
  font-size: 1.05rem; letter-spacing: .04em; color: #fbe6d3; margin-right: 40px; white-space: nowrap;
}
.tabs > .tab-nav > button { color: #e8dbe6 !important; font-weight: 600 !important; padding: 8px 18px !important; }
.tabs > .tab-nav > button.selected {
  background: linear-gradient(135deg, var(--amber), var(--coral)) !important; color: #ffffff !important;
}

button.primary {
  border: none !important; font-weight: 700 !important; border-radius: 999px !important;
  box-shadow: 0 6px 18px rgba(226,114,91,.28);
  transition: transform .15s ease, box-shadow .15s ease;
}
button.primary:hover { transform: translateY(-2px); box-shadow: 0 10px 24px rgba(226,114,91,.36); }
button.secondary {
  background: #ffffff !important;
  border: 2px solid var(--coral) !important;
  color: var(--coral) !important;
  border-radius: 999px !important; font-weight: 700 !important;
  box-shadow: 0 4px 14px rgba(226,114,91,.15);
  transition: transform .15s ease, background .15s ease, color .15s ease;
}
button.secondary:hover { transform: translateY(-2px); background: var(--coral) !important; color: #ffffff !important; }
.block { border-radius: 18px !important; transition: box-shadow .2s ease, transform .2s ease; }
.gr-box, .form { border-radius: 18px !important; }
input[type="number"], input[type="text"], textarea, select {
  background: #ffffff !important;
  border: 1.5px solid #d9c9a3 !important;
  box-shadow: inset 0 1px 3px rgba(180,120,40,.06) !important;
}
input[type="number"]:focus, input[type="text"]:focus, textarea:focus, select:focus {
  border-color: var(--coral) !important;
  box-shadow: 0 0 0 3px rgba(226,114,91,.15) !important;
}
table { font-family: 'JetBrains Mono', monospace !important; font-size: .82rem !important; }
thead th { background: linear-gradient(135deg, #fdf1de, #fbe6d3) !important; color: #b5591a !important; }
.prose a { color: var(--coral) !important; font-weight: 600; }
footer { opacity: .55; }

/* hero (Home tab) */
/* hero (Home tab) */
.hero-banner { text-align: center; padding: 10px 20px 6px; }
.hero-banner .hero-mark {
  font-family: 'Fraunces', serif; font-size: 3.6rem; font-weight: 700; letter-spacing: .1em; margin: 0 0 10px;
  color: var(--coral); /* fallback if gradient-text isn't supported */
  background: linear-gradient(135deg, var(--coral), var(--amber), var(--violet));
  background-clip: text; -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
  color: transparent;
}
@supports not (background-clip: text) {
  .hero-banner .hero-mark { color: var(--coral); -webkit-text-fill-color: unset; }
}
.hero-banner .hero-sub { font-family: 'Fraunces', serif; font-style: italic; font-size: 1.2rem; color: #6b5c46; margin-bottom: 18px; }
.hero-banner .hero-strip {
  display: inline-block; background: #fdf1de; border: 1px solid #f0d9a8; color: #b5591a;
  font-family: 'Fraunces', serif; font-weight: 600; font-size: 1.5rem; padding: 14px 34px;
  border-radius: 999px; letter-spacing: .01em;
}
.hero-welcome { max-width: 780px; margin: 22px auto 0; text-align: justify; text-justify: inter-word; font-size: 1.02rem; line-height: 1.75; color: #4a4232; }
.hero-welcome .eyebrow { text-align: center; font-family: 'JetBrains Mono', monospace; letter-spacing: .25em; font-size: 1.05rem; font-weight: 600; color: #a89a7a; text-transform: uppercase; margin-bottom: 16px; }
.hero-welcome .tagline { font-family: 'Fraunces', serif; font-weight: 700; font-size: 1.4rem; color: var(--coral); margin-top: 22px; }

/* developer cards — restyled to match the light theme */
.dev-wrap .dev-card {
  background: linear-gradient(145deg, #ffffff, #fdf3e2) !important;
  border: 1px solid #ecdfc4 !important;
}
.dev-wrap .dev-name { color: var(--ink) !important; }
.dev-wrap .dev-role { color: var(--coral) !important; }
.dev-wrap .dev-affil { color: #8a7f68 !important; }
.dev-wrap .dev-intro { color: #4a4232 !important; }
.dev-wrap .dev-photo-wrap { background: linear-gradient(135deg, var(--amber), var(--violet)) !important; }
.dev-wrap .dev-btn-email { background: var(--teal) !important; }
.dev-wrap .dev-btn-email:hover { background: #24a396 !important; }
.dev-wrap .dev-btn-scholar { background: var(--violet) !important; }
.dev-wrap .dev-btn-scholar:hover { background: #7c4ee0 !important; }
.dev-wrap .dev-card .dev-links a.dev-btn {
  color: #ffffff !important;
  text-shadow: 0 1px 1px rgba(0,0,0,.2);
}

.dev-wrap .dev-grid {
  display: grid !important;
  grid-template-columns: repeat(2, 1fr);
  gap: 16px !important;
}
.dev-wrap { max-width: 900px !important; }
@media (max-width: 700px) {
  .dev-wrap .dev-grid { grid-template-columns: 1fr; }
}
/* ── Instructions page — card layout ── */
.instr-section { margin-bottom: 34px; }
.instr-section h3 {
  font-family: 'Fraunces', serif; font-size: 1.3rem; color: var(--ink);
  display: flex; align-items: center; gap: 10px; margin-bottom: 18px;
}
.instr-section h3::before {
  content: ''; width: 26px; height: 2px; background: linear-gradient(90deg, var(--coral), var(--amber));
}
.step-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); gap: 14px; }
.step-card {
  background: #ffffff; border: 1px solid #ecdfc4; border-radius: 16px; padding: 18px 20px;
  box-shadow: 0 4px 16px rgba(180,120,40,.08);
  transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease;
}
.step-card:hover { transform: translateY(-3px); box-shadow: 0 10px 26px rgba(226,114,91,.18); border-color: var(--coral); }
.step-card .step-num {
  display: inline-flex; align-items: center; justify-content: center;
  width: 30px; height: 30px; border-radius: 999px; margin-bottom: 10px;
  background: linear-gradient(135deg, var(--amber), var(--coral)); color: #fff;
  font-family: 'JetBrains Mono', monospace; font-weight: 700; font-size: .85rem;
}
.step-card .step-title { font-family: 'Fraunces', serif; font-weight: 600; font-size: 1rem; color: var(--ink); margin-bottom: 6px; }
.step-card .step-body { font-size: .87rem; color: #6b5c46; line-height: 1.5; }
.formula-list { display: flex; flex-direction: column; gap: 10px; }
.formula-line {
  background: #ffffff; border: 1px solid #ecdfc4; border-radius: 12px; padding: 12px 16px;
  display: flex; align-items: center; gap: 14px; font-size: .89rem; color: #4a4232;
  transition: border-color .18s ease, transform .18s ease;
}
.formula-line:hover { border-color: var(--teal); transform: translateX(4px); }
.formula-line .f-num {
  font-family: 'JetBrains Mono', monospace; font-weight: 700; color: var(--teal); flex-shrink: 0; width: 18px;
}
.formula-line code {
  background: #fdf1de; color: #b5591a; padding: 3px 9px; border-radius: 7px;
  font-family: 'JetBrains Mono', monospace; font-size: .84rem;
}
.note-box {
  background: #fdf1de; border: 1px dashed #f0d9a8; border-radius: 14px; padding: 16px 20px;
  font-style: italic; color: #8a7f68; font-size: .88rem;
}
.methodology-box {
  background: #ffffff; border: 1px solid #ecdfc4; border-radius: 18px;
  padding: 28px 32px; box-shadow: 0 4px 18px rgba(180,120,40,.08);
  position: relative;
}
.methodology-box::before {
  content: ''; position: absolute; top: 0; left: 0; width: 5px; height: 100%;
  background: linear-gradient(180deg, var(--coral), var(--amber));
  border-radius: 18px 0 0 18px;
}
.methodology-box p {
  margin: 0 0 14px; padding-left: 14px; color: #4a4232; font-size: .98rem;
  line-height: 1.9; text-align: justify;
}
.methodology-box p:last-child { margin-bottom: 0; }
.formula-block {
  background: #fdf1de; border: 1px solid #f0d9a8; border-radius: 12px;
  padding: 18px 20px; margin: 4px 14px 16px; text-align: center;
  font-family: 'JetBrains Mono', monospace; font-size: 1.05rem; color: #b5591a;
  font-style: italic; overflow-x: auto;
}
.formula-block--small { font-size: .95rem; }
.sum { font-size: 1.3em; margin: 0 2px; }
.site-footer {
  text-align: center; margin-top: 40px; padding: 18px 0;
  border-top: 1px solid #ecdfc4; color: #8a7f68; font-size: .82rem; line-height: 1.7;
}
"""

with gr.Blocks(title="GARCH-Guided LSTM (GGLSTM)", theme=CUSTOM_THEME, css=CUSTOM_CSS, head=FONT_HEAD) as demo:
    df_state = gr.State(None)
    ar_garch_state = gr.State(None)
    best_hp_state = gr.State(None)

    gr.HTML(f'<div id="nav-title"><img src="{_img_to_data_uri("icar_logo.jpg")}" alt="ICAR-IASRI"/></div>')

    with gr.Tabs():
        with gr.Tab("Home"):
            gr.HTML(
                """
                <div class="hero-banner">
                  <div class="hero-mark">G A R C H · L S T M</div>
                  <span class="hero-strip">GARCH-Guided LSTM (GGLSTM)</span>
                </div>
                """
            )
            gr.HTML(
                """
                <div class="hero-welcome">
                  <div class="eyebrow">— Welcome —</div>
                  <p>The <strong>GARCH-Guided Long Short-Term Memory (GG-LSTM)</strong> model is a
                  statistically informed forecasting framework inspired by the principles of
                  <strong>Physics-Informed Machine Learning (PIML)</strong>, wherein established domain
                  knowledge is incorporated into the learning process to improve model generalization
                  and predictive performance. It integrates econometric volatility modelling with
                  deep learning-based nonlinear sequence learning for improved volatility forecasting.
                  The framework combines conditional variance information estimated from the GARCH
                  model with the temporal feature extraction capability of LSTM networks. Unlike
                  conventional hybrid approaches that use GARCH outputs only as input variables,
                  GG-LSTM incorporates GARCH-derived volatility information directly into the
                  optimization process through a composite loss function. The proposed loss function
                  balances forecasting accuracy based on realized variance and consistency with
                  GARCH-estimated conditional volatility dynamics. This integration enables the model
                  to capture both volatility persistence and complex nonlinear market behaviour. The
                  proposed framework provides a statistically informed approach for volatility
                  forecasting.</p>
                </div>
                """
            )

        with gr.Tab("Model"):
            with gr.Tabs() as model_tabs:
                with gr.Tab("1. Data", id=0):
                    gr.Markdown("Upload a CSV or Excel file containing a **time** column and a **study variable** column.")
                    file_in = gr.File(label="Upload Data", file_types=[".csv", ".xlsx", ".xls"])
                    load_msg = gr.Markdown(value="Upload a file above to begin.", min_height=30)
                    with gr.Row():
                        time_col = gr.Dropdown(label="Time Column", choices=[])
                        value_col = gr.Dropdown(label="Study Variable", choices=[])
                    file_in.change(load_file, inputs=file_in, outputs=[df_state, time_col, value_col, load_msg])
                    gr.Examples(
                        examples=[[os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data.csv")]],
                        inputs=file_in, outputs=[df_state, time_col, value_col, load_msg],
                        fn=load_file, cache_examples=False, label="Or try it with sample data",
                    )
                    next_to_stats_btn = gr.Button("Next → Summary Statistics", variant="secondary")
                    next_to_stats_btn.click(lambda: gr.Tabs(selected=1), outputs=model_tabs)

                with gr.Tab("2. Summary Statistics", id=1):
                    select_all_stats = gr.Checkbox(label="Select All", value=False)
                    stats_pick = gr.CheckboxGroup(STAT_CHOICES, value=["Mean", "Median", "Standard Deviation"],
                                                   label="Statistics to display")

                    def _toggle_all_stats(checked):
                        return gr.update(value=STAT_CHOICES if checked else [])

                    select_all_stats.change(_toggle_all_stats, inputs=select_all_stats, outputs=stats_pick)
                    with gr.Row():
                        show_line = gr.Checkbox(label="Show Line Plot", value=True)
                        show_box = gr.Checkbox(label="Show Boxplot", value=True)
                    with gr.Accordion("Plot Size Adjuster", open=False):
                        with gr.Row():
                            plot_width_in = gr.Slider(4, 16, value=8, step=0.5, label="Plot Width")
                            plot_height_in = gr.Slider(2.5, 10, value=3.5, step=0.5, label="Plot Height")
                            font_size_in = gr.Slider(6, 20, value=9, step=1, label="Axis Number Font Size")
                    stats_btn = gr.Button("Compute", variant="primary")
                    stats_out = gr.Markdown(value="Click **Compute** to see statistics here.", min_height=30)
                    line_plot = gr.Plot(label="Time Series")
                    box_plot = gr.Plot(label="Boxplot")
                    stats_btn.click(summary_stats,
                                     inputs=[df_state, time_col, value_col, stats_pick, show_line, show_box,
                                             plot_width_in, plot_height_in, font_size_in],
                                     outputs=[stats_out, line_plot, box_plot])
                    next_to_garch_btn = gr.Button("Next → AR & GARCH", variant="secondary")
                    next_to_garch_btn.click(lambda: gr.Tabs(selected=2), outputs=model_tabs)


                with gr.Tab("3. AR & GARCH", id=2):
                    gr.Markdown("Auto-tunes an AR model for the mean return and a GARCH model "
                                "(grid over p,q ∈ {1,2,3}, mean ∈ {Zero,AR}, dist ∈ {normal,t} — 36 "
                                "combinations) for the variance, selecting the lowest-AIC fit for each.")
                    ar_train_ratio = gr.Slider(0.5, 0.95, value=0.8, step=0.05, label="Train / Test Split Ratio")
                    ar_garch_btn = gr.Button("Fit AR & GARCH", variant="primary")
                    ar_garch_out = gr.Markdown(value="Click **Fit AR & GARCH** to see results here.", min_height=40)
                    ar_garch_plot = gr.Plot()
                    gr.Markdown("**GARCH Parameter Estimates**")
                    garch_params_html = gr.HTML()
                    ar_garch_download = gr.File(label="Download AR & GARCH Results (Excel)")
                    ar_garch_btn.click(run_ar_garch,
                                        inputs=[df_state, value_col, ar_train_ratio],
                                        outputs=[ar_garch_out, ar_garch_plot, ar_garch_state, garch_params_html, ar_garch_download])
                    next_to_lstm_btn = gr.Button("Next → LSTM Hyperparameters", variant="secondary")
                    next_to_lstm_btn.click(lambda: gr.Tabs(selected=3), outputs=model_tabs)

                with gr.Tab("4. LSTM Hyperparameters", id=3):
                    seq_len = gr.Slider(2, 30, value=5, step=1, label="Sequence Length (lag window)")
                    tune_mode = gr.Radio(["Manual", "Grid Search", "Bayesian Search"],
                                          value="Manual", label="Hyperparameter Selection")

                    with gr.Group(visible=True) as manual_group:
                        gr.Markdown("**Manual** — set exact hyperparameter values.")
                        with gr.Row():
                            hidden_in = gr.Number(label="Hidden Size", value=32, precision=0)
                            layers_in = gr.Number(label="Num Layers", value=1, precision=0)
                            lr_in = gr.Number(label="Learning Rate", value=0.003)
                            dropout_in = gr.Number(label="Dropout", value=0.1)

                    with gr.Group(visible=False) as grid_group:
                        gr.Markdown("**Grid Search** — set a min, max, and step for each hyperparameter.")
                        with gr.Row():
                            g_hidden_min = gr.Number(label="Hidden Size — Min", value=8, precision=0)
                            g_hidden_max = gr.Number(label="Hidden Size — Max", value=40, precision=0)
                            g_hidden_step = gr.Number(label="Hidden Size — Step", value=8, precision=0)
                        with gr.Row():
                            g_layers_min = gr.Number(label="Num Layers — Min", value=1, precision=0)
                            g_layers_max = gr.Number(label="Num Layers — Max", value=2, precision=0)
                            g_layers_step = gr.Number(label="Num Layers — Step", value=1, precision=0)
                        with gr.Row():
                            g_lr_min = gr.Number(label="Learning Rate — Min", value=0.001)
                            g_lr_max = gr.Number(label="Learning Rate — Max", value=0.005)
                            g_lr_step = gr.Number(label="Learning Rate — Step", value=0.002)
                        with gr.Row():
                            g_dropout_min = gr.Number(label="Dropout — Min", value=0.0)
                            g_dropout_max = gr.Number(label="Dropout — Max", value=0.3)
                            g_dropout_step = gr.Number(label="Dropout — Step", value=0.1)

                    with gr.Group(visible=False) as bayes_group:
                        gr.Markdown("**Bayesian Search** — Gaussian Process optimization (scikit-optimize).")
                        with gr.Row():
                            b_hidden_min = gr.Number(label="Hidden Size — Min", value=8, precision=0)
                            b_hidden_max = gr.Number(label="Hidden Size — Max", value=40, precision=0)
                        with gr.Row():
                            b_layers_min = gr.Number(label="Num Layers — Min", value=1, precision=0)
                            b_layers_max = gr.Number(label="Num Layers — Max", value=2, precision=0)
                        with gr.Row():
                            b_lr_min = gr.Number(label="Learning Rate — Min", value=0.0005)
                            b_lr_max = gr.Number(label="Learning Rate — Max", value=0.006)
                        with gr.Row():
                            b_dropout_min = gr.Number(label="Dropout — Min", value=0.0)
                            b_dropout_max = gr.Number(label="Dropout — Max", value=0.3)
                        n_calls_in = gr.Slider(5, 40, value=15, step=1, label="Search Iterations")

                    def _toggle_tune_panels(mode):
                        return (gr.update(visible=mode == "Manual"),
                                gr.update(visible=mode == "Grid Search"),
                                gr.update(visible=mode == "Bayesian Search"))

                    tune_mode.change(_toggle_tune_panels, inputs=tune_mode,
                                      outputs=[manual_group, grid_group, bayes_group])

                    lstm_btn = gr.Button("Set / Tune Hyperparameters", variant="primary")
                    lstm_out = gr.Markdown(value="Click **Set / Tune Hyperparameters** to begin.", min_height=60)
                    lstm_btn.click(run_lstm_tuning,
                                    inputs=[ar_garch_state, seq_len, tune_mode,
                                            hidden_in, layers_in, lr_in, dropout_in,
                                            g_hidden_min, g_hidden_max, g_hidden_step,
                                            g_layers_min, g_layers_max, g_layers_step,
                                            g_lr_min, g_lr_max, g_lr_step,
                                            g_dropout_min, g_dropout_max, g_dropout_step,
                                            b_hidden_min, b_hidden_max,
                                            b_layers_min, b_layers_max,
                                            b_lr_min, b_lr_max,
                                            b_dropout_min, b_dropout_max,
                                            n_calls_in],
                                    outputs=[lstm_out, best_hp_state])
                    next_to_ginn_btn = gr.Button("Next → GINN", variant="secondary")
                    next_to_ginn_btn.click(lambda: gr.Tabs(selected=4), outputs=model_tabs)

                with gr.Tab("5. GGLSTM", id=4):
                    gr.Markdown("Runs the LSTM for **λ = 0.0 (Standard, pure ground truth)** plus each λ "
                                "you list below (λ=1.0 is pure GARCH-matching, i.e. GGLSTM-0), using the "
                                "AR/GARCH fit and LSTM hyperparameters set in the previous steps.")
                    lambdas_in = gr.Textbox(label="λ values (comma-separated)",
                                             value="0.0, 0.1, 0.3, 0.5, 0.7, 0.9")
                    with gr.Row():
                        epochs_in = gr.Number(label="Max Epochs", value=500, precision=0)
                        patience_in = gr.Number(label="Early-Stopping Patience", value=30, precision=0)
                    run_btn = gr.Button("Run GGLSTM", variant="primary")
                    run_status = gr.Markdown(value="Click **Run GGLSTM** to begin.", min_height=60)
                    gr.Markdown("### Configuration & Hyperparameters Used")
                    config_table = gr.Dataframe(label="Configuration", headers=["Parameter", "Value"])
                    gr.Markdown("### Results (Train & Test Metrics)")
                    results_html = gr.HTML()
                    gr.Markdown("### Actual vs. Predicted — by λ")
                    with gr.Accordion("Training Set: Actual vs Predicted", open=False):
                        train_pred_html = gr.HTML()
                    with gr.Accordion("Test Set: Actual vs Predicted", open=True):
                        test_pred_html = gr.HTML()
                    download_file = gr.File(label="Download Excel Results")
                    run_btn.click(run_ginn_pipeline,
                                  inputs=[ar_garch_state, seq_len, best_hp_state, lambdas_in, epochs_in, patience_in],
                                  outputs=[run_status, results_html, download_file,
                                           train_pred_html, test_pred_html, config_table])

        with gr.Tab("Instructions"):
            gr.HTML(
                """
                <div class="instr-section">
                  <h3>Methodology</h3>
                  <div class="methodology-box">
                    <p>The proposed GARCH-Guided Long Short-Term Memory (GGLSTM) framework integrates
                    econometric volatility modelling with deep learning-based nonlinear sequence
                    learning for onion price volatility forecasting. The observed onion price series
                    was transformed into logarithmic returns to obtain a stationary representation of
                    price fluctuations. Since the true conditional variance is unobservable, realized
                    variance was calculated from observed returns and considered as the empirical
                    volatility target. A GARCH model was then employed to estimate conditional
                    variance dynamics. The LSTM network was trained to forecast future realized
                    variance. To incorporate econometric volatility information during model
                    optimization, a GARCH-guided composite loss function was developed:</p>

                    <div class="formula-block">
                      L<sub>GG&minus;LSTM</sub> &nbsp;=&nbsp; (1&nbsp;&minus;&nbsp;λ)&nbsp;L<sub>RV</sub> &nbsp;+&nbsp; λL<sub>GARCH</sub>
                    </div>

                    <p>where:</p>
                    <div class="formula-block formula-block--small">
                      L<sub>RV</sub> &nbsp;=&nbsp; <span class="sum">Σ</span>
                      (RV<sub>t</sub> &minus; σ̂²<sub>L,t</sub>)<sup>2</sup>
                    </div>
                    <p>represents the realized variance forecasting error, and:</p>
                    <div class="formula-block formula-block--small">
                      L<sub>GARCH</sub> &nbsp;=&nbsp; <span class="sum">Σ</span>
                      (σ̂²<sub>G,t</sub> &minus; σ̂²<sub>L,t</sub>)<sup>2</sup>
                    </div>
                    <p>represents the deviation between LSTM predictions and GARCH-estimated conditional
                    variance. The parameter λ controls the relative contribution of observed volatility
                    information and GARCH-based volatility dynamics during training.</p>
                  </div>
                </div>

                <div class="instr-section">
                  <h3>How to use this tool</h3>
                  <div class="step-grid">
                    <div class="step-card">
                      <div class="step-num">1</div>
                      <div class="step-title">Data</div>
                      <div class="step-body">Upload a CSV/Excel file, then pick the time column and the study variable (a price, yield, or other series with meaningful returns).</div>
                    </div>
                    <div class="step-card">
                      <div class="step-num">2</div>
                      <div class="step-title">Summary Statistics</div>
                      <div class="step-body">Choose which statistics and plots (line, boxplot) to view.</div>
                    </div>
                    <div class="step-card">
                      <div class="step-num">3</div>
                      <div class="step-title">AR &amp; GARCH</div>
                      <div class="step-body">Set the train/test split, then fit — auto-tunes the AR mean model and searches GARCH configurations by AIC.</div>
                    </div>
                    <div class="step-card">
                      <div class="step-num">4</div>
                      <div class="step-title">LSTM Hyperparameters</div>
                      <div class="step-body">Choose manual values, or auto-tune with grid search or Bayesian search (Gaussian Process).</div>
                    </div>
                    <div class="step-card">
                      <div class="step-num">5</div>
                      <div class="step-title">GGLSTM</div>
                      <div class="step-body">List the λ values to test (0 = Standard LSTM, 1 = pure GARCH-matching) — trains one model per λ plus the Standard baseline, reporting RMSE / MAE / SMAPE, with a downloadable Excel workbook.</div>
                    </div>
                  </div>
                </div>

                <div class="instr-section">
                  <h3>Pipeline</h3>
                  <div class="formula-list">
                    <div class="formula-line"><span class="f-num">1</span> Compute returns: <code>y_t → r_t = (y_t − y_{t−1}) / y_{t−1}</code></div>
                    <div class="formula-line"><span class="f-num">2</span> AR(p) → mean forecast μ̂_t &nbsp;·&nbsp; GARCH(p,q) → variance forecast σ²̂_GARCH</div>
                    <div class="formula-line"><span class="f-num">3</span> Ground truth variance: <code>σ²_t = (r_t − μ̂_t)²</code></div>
                    <div class="formula-line"><span class="f-num">4</span> LSTM learns to predict σ²_t from past σ²_t sequences</div>
                    <div class="formula-line"><span class="f-num">5</span> Loss = (1−λ)·MSE(σ²_t, σ²̂_LSTM) + λ·MSE(σ²̂_GARCH, σ²̂_LSTM)</div>
                  </div>
                </div>
                """
            )
        with gr.Tab("Developers"):
            gr.Markdown("## Developers")
            gr.HTML(build_developers_html())

    footer_html = gr.HTML()
    demo.load(
        lambda: (f"<div class='site-footer'>"
                 f"Copyright © 2026 ICAR – Indian Agricultural Statistics Research Institute, "
                 f"New Delhi - 110012. All Rights Reserved.<br>"
                 f"👥 Total Visitors: {get_and_increment_visitor_count()}</div>"),
        outputs=footer_html,
    )

if __name__ == "__main__":
    import os
    demo.queue().launch(
        server_name="0.0.0.0",
        server_port=int(os.environ.get("PORT", 7860)),
        share=True,
        show_api=False
    )